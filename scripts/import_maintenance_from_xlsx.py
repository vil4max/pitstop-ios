#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Downloads" / "VW ТО-2.xlsx"
OUT_PLAN = REPO / "Pitstop" / "Resources" / "maintenance-plan.json"
OUT_UPGRADES = REPO / "Pitstop" / "Resources" / "upgrades.json"

PURCHASE_DATE = "2024-05-22"
CURRENT_ODOMETER = 13600
MILEAGE_BASELINE_KM = 13500
MILEAGE_BASELINE_MONTHS = 24
INTERVAL_KM_MIN = 5000
INTERVAL_KM_MAX = 7000
INTERVAL_KM_DEFAULT = 6000
MILESTONE_FROM_KM = 22500
SCHEMA_VERSION = 4
DEALER_NAME = "AVTODIM ATLANT"
EXPECTED_OIL_CHANGE_COST_UAH = 3000

COMPLETED_REGULAR = [
    {
        "sortOrder": 3000,
        "targetOdometerKm": 3000,
        "completedOdometer": 3000,
        "completedAt": "2024-11-14",
        "serviceScope": "seasonal",
        "includesOilChange": False,
        "costProfile": "seasonalNoOil",
        "odometerIsEstimate": True,
        "dealer": DEALER_NAME,
        "costUah": 9940,
        "payments": [
            {"date": "2024-11-10", "amountUah": 5876, "card": "platinum", "payee": DEALER_NAME},
            {"date": "2024-11-14", "amountUah": 4064, "card": "platinum", "payee": DEALER_NAME},
        ],
    },
    {
        "sortOrder": 6000,
        "targetOdometerKm": 6000,
        "completedOdometer": 6000,
        "completedAt": "2025-05-01",
        "serviceScope": "oilService",
        "includesOilChange": True,
        "costProfile": "dealerPackage",
        "estimatedOilPortionUah": EXPECTED_OIL_CHANGE_COST_UAH,
        "odometerIsEstimate": False,
        "dealer": DEALER_NAME,
        "costUah": 17946,
        "payments": [
            {
                "date": "2025-05-01",
                "amountUah": 17946,
                "card": "white",
                "payee": DEALER_NAME,
            }
        ],
    },
    {
        "sortOrder": 11000,
        "targetOdometerKm": 11000,
        "completedOdometer": 11000,
        "completedAt": "2026-03-13",
        "serviceScope": "oilService",
        "includesOilChange": True,
        "costProfile": "majorService",
        "estimatedOilPortionUah": EXPECTED_OIL_CHANGE_COST_UAH,
        "odometerIsEstimate": False,
        "dealer": DEALER_NAME,
        "costUah": 27764,
        "payments": [
            {
                "date": "2026-03-13",
                "amountUah": 27764,
                "card": "platinum",
                "payee": DEALER_NAME,
            }
        ],
    },
]

DEALER_OTHER_PAYMENTS = [
    {
        "date": "2026-04-02",
        "amountUah": 869,
        "card": "platinum",
        "payee": "ТОВ 'Автомобільний ДІМ Атлант'",
        "category": "softwareSetup",
        "title": "Налаштування ПО",
    },
]

DEALER_PAYMENTS_EXCLUDED = [
    {
        "date": "2024-03-20",
        "amountUah": 15473,
        "card": "white",
        "payee": DEALER_NAME,
        "reason": "beforeVehiclePurchase",
    },
    {
        "date": "2024-04-13",
        "amountUah": 10983,
        "card": "platinum",
        "payee": DEALER_NAME,
        "reason": "beforeVehiclePurchase",
    },
]


OIL_TASK_TITLES = {
    "замена масла в двигателе",
    "масляный фильтр",
    "уплотнительное кольцо",
}


def format_km_part(km: int) -> str:
    return f"{km:,}".replace(",", ".")


def format_km_title(km: int) -> str:
    return f"{format_km_part(km)} km"


def format_interval_title(from_km: int, to_km: int) -> str:
    return f"{format_km_part(from_km)}–{format_km_part(to_km)} km"


def average_monthly_km(distance_km: int, months: int) -> int:
    return int(distance_km / months + 0.5)


def parse_km(raw: str) -> int | None:
    if not raw:
        return None
    text = raw.strip()
    low = text.lower()
    if "каждые" in low or "год" in low or "моточас" in low:
        return None
    match = re.search(r"([\d.,]+)", text)
    if not match:
        return None
    val = match.group(1).replace(",", ".")
    num = float(val)
    if "." in val and num < 1000:
        return int(num * 1000)
    if "тыс" in low:
        return int(num * 1000)
    return int(num)


def load_xlsx_visits(xlsx: Path) -> tuple[list[dict], dict[str, bool]]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Обслуживание"]

    default_mandatory: dict[str, bool] = {}
    for row in range(2, 8):
        title = ws.cell(row, 2).value
        flag = ws.cell(row, 3).value
        if title:
            default_mandatory[str(title).strip().lower()] = str(flag).lower() == "y"

    visits: list[dict] = []
    current: dict | None = None
    in_planned = False

    for row in range(1, ws.max_row + 1):
        header = ws.cell(row, 1).value
        title = ws.cell(row, 2).value
        mandatory_flag = ws.cell(row, 3).value

        if header and isinstance(header, str) and "Плановое" in header:
            in_planned = True
            continue
        if not in_planned:
            continue

        if header:
            km = parse_km(str(header))
            if km is None:
                continue
            current = {"targetOdometerKm": km, "tasks": []}
            visits.append(current)

        if title and current:
            task_title = str(title).strip()
            if mandatory_flag is not None:
                is_mandatory = str(mandatory_flag).strip().lower() == "y"
            else:
                is_mandatory = default_mandatory.get(task_title.lower(), False)
            current["tasks"].append({"title": task_title, "isMandatory": is_mandatory})

    return visits, default_mandatory


def load_milestones_from_plan(plan_path: Path) -> list[dict]:
    if not plan_path.exists():
        return []
    plan = json.loads(plan_path.read_text(encoding="utf-8"))
    milestones: list[dict] = []
    for visit in plan.get("visits", []):
        if visit.get("kind") != "milestone":
            continue
        milestones.append(
            {
                "targetOdometerKm": visit["targetOdometerKm"],
                "tasks": [
                    {"title": task["title"], "isMandatory": task["isMandatory"]}
                    for task in visit["tasks"]
                ],
            }
        )
    milestones.sort(key=lambda item: item["targetOdometerKm"])
    return milestones


def is_oil_task(title: str) -> bool:
    return title.strip().lower() in OIL_TASK_TITLES


def tasks_to_json(
    tasks: list[dict],
    visit_id: str,
    *,
    completed: bool,
    includes_oil_change: bool,
) -> list[dict]:
    result = []
    for index, task in enumerate(tasks, start=1):
        oil = is_oil_task(task["title"])
        if includes_oil_change:
            is_applicable = True
            is_mandatory = task["isMandatory"]
            is_done = completed
        else:
            is_applicable = not oil
            is_mandatory = task["isMandatory"] if is_applicable else False
            is_done = completed and is_applicable
        result.append(
            {
                "id": f"{visit_id}-t{index}",
                "title": task["title"],
                "sortOrder": index,
                "isMandatory": is_mandatory,
                "isApplicable": is_applicable,
                "isDone": is_done,
            }
        )
    return result


def merge_task_lists(base_tasks: list[dict], extra_tasks: list[dict]) -> list[dict]:
    merged = list(base_tasks)
    seen = {task["title"].strip().lower() for task in merged}
    has_oil = any(is_oil_task(task["title"]) for task in merged)
    for task in extra_tasks:
        title = task["title"].strip()
        key = title.lower()
        if key in seen:
            continue
        if has_oil and is_oil_task(title):
            continue
        merged.append({"title": title, "isMandatory": task["isMandatory"]})
        seen.add(key)
    return merged


def generate_interval_windows(last_oil_km: int, max_milestone_km: int) -> list[tuple[int, int, int]]:
    windows: list[tuple[int, int, int]] = []
    anchor = last_oil_km
    while True:
        window_from = anchor + INTERVAL_KM_MIN
        window_to = anchor + INTERVAL_KM_MAX
        target = anchor + INTERVAL_KM_DEFAULT
        if window_from > max_milestone_km:
            break
        windows.append((window_from, target, window_to))
        anchor = window_to
    return windows


def assign_milestones_to_windows(
    milestones: list[dict],
    windows: list[tuple[int, int, int]],
) -> list[list[dict]]:
    buckets: list[list[dict]] = [[] for _ in windows]
    for milestone in milestones:
        km = milestone["targetOdometerKm"]
        chosen_index = None
        for index, (window_from, _, window_to) in enumerate(windows):
            if window_from <= km <= window_to:
                chosen_index = index
                break
        if chosen_index is None:
            chosen_index = min(
                range(len(windows)),
                key=lambda index: abs(km - windows[index][1]),
            )
        buckets[chosen_index].append(milestone)
    return buckets


def build_regular_visit(
    sort_km: int,
    tasks: list[dict],
    *,
    completed: bool,
    target_odometer_km: int | None = None,
    completed_odometer: int | None = None,
    completed_at: str | None = None,
    title: str | None = None,
    dealer: str | None = None,
    cost_uah: int | None = None,
    payments: list[dict] | None = None,
    includes_oil_change: bool = True,
    service_scope: str = "oilService",
    odometer_is_estimate: bool = False,
    cost_profile: str | None = None,
    estimated_oil_portion_uah: int | None = None,
    window_from_km: int | None = None,
    window_to_km: int | None = None,
) -> dict:
    target = target_odometer_km if target_odometer_km is not None else sort_km
    visit_id = f"visit-{sort_km}"
    visit_title = title if title is not None else format_km_title(target)
    visit: dict = {
        "id": visit_id,
        "title": visit_title,
        "kind": "regular",
        "sortOrder": sort_km,
        "targetOdometerKm": target,
        "serviceScope": service_scope,
        "includesOilChange": includes_oil_change,
        "tasks": tasks_to_json(
            tasks,
            visit_id,
            completed=completed,
            includes_oil_change=includes_oil_change,
        ),
        "isCompleted": completed,
    }
    if completed:
        visit["completedOdometer"] = completed_odometer if completed_odometer is not None else target
        if completed_at:
            visit["completedAt"] = completed_at
        if dealer:
            visit["dealer"] = dealer
        if cost_uah is not None:
            visit["costUah"] = cost_uah
        if cost_profile:
            visit["costProfile"] = cost_profile
        if estimated_oil_portion_uah is not None:
            visit["estimatedOilPortionUah"] = estimated_oil_portion_uah
        if payments:
            visit["payments"] = payments
        if odometer_is_estimate:
            visit["odometerIsEstimate"] = True
    if window_from_km is not None and window_to_km is not None:
        visit["windowFromKm"] = window_from_km
        visit["windowToKm"] = window_to_km
        visit["windowKm"] = max((window_to_km - target), (target - window_from_km))
    return visit


def build_interval_visit(
    window_from: int,
    target: int,
    window_to: int,
    base_tasks: list[dict],
    milestone_buckets: list[dict],
) -> dict:
    merged_tasks = merge_task_lists(base_tasks, [task for milestone in milestone_buckets for task in milestone["tasks"]])
    return build_regular_visit(
        target,
        merged_tasks,
        completed=False,
        target_odometer_km=target,
        title=format_interval_title(window_from, window_to),
        includes_oil_change=True,
        service_scope="oilService",
        window_from_km=window_from,
        window_to_km=window_to,
    )


def build_dealer_statement() -> dict:
    vehicle_visits = []
    total_uah = 0
    for entry in COMPLETED_REGULAR:
        total_uah += entry["costUah"]
        vehicle_visits.append(
            {
                "visitId": f"visit-{entry['sortOrder']}",
                "serviceDate": entry["completedAt"],
                "completedOdometerKm": entry["completedOdometer"],
                "odometerIsEstimate": entry.get("odometerIsEstimate", False),
                "includesOilChange": entry["includesOilChange"],
                "costProfile": entry.get("costProfile"),
                "estimatedOilPortionUah": entry.get("estimatedOilPortionUah"),
                "totalUah": entry["costUah"],
                "payments": entry.get("payments", []),
            }
        )
    return {
        "dealer": DEALER_NAME,
        "source": "Monobank search «atl»",
        "purchaseDate": PURCHASE_DATE,
        "vehicleVisitsTotalUah": total_uah,
        "vehicleVisits": vehicle_visits,
        "otherPayments": DEALER_OTHER_PAYMENTS,
        "otherPaymentsTotalUah": sum(item["amountUah"] for item in DEALER_OTHER_PAYMENTS),
        "excludedPayments": DEALER_PAYMENTS_EXCLUDED,
        "excludedNote": "Before Arteon purchase (22 May 2024) — previous car or prepurchase, not in app history.",
    }


def import_plan(xlsx: Path | None, odometer: int) -> dict:
    if xlsx is not None and xlsx.exists():
        xlsx_visits, _ = load_xlsx_visits(xlsx)
        regular_template = next(v for v in xlsx_visits if v["targetOdometerKm"] == 7500)["tasks"]
        milestones = [v for v in xlsx_visits if v["targetOdometerKm"] >= MILESTONE_FROM_KM]
    else:
        regular_template = [
            {"title": "Замена Масла в Двигателе", "isMandatory": True},
            {"title": "Масляный фильтр", "isMandatory": True},
            {"title": "Уплотнительное кольцо", "isMandatory": True},
            {"title": "Воздушный Фильтр", "isMandatory": True},
            {"title": "Салонный фильтр", "isMandatory": True},
            {"title": "Топливная присадка G17", "isMandatory": False},
        ]
        milestones = load_milestones_from_plan(OUT_PLAN)

    avg_monthly = average_monthly_km(MILEAGE_BASELINE_KM, MILEAGE_BASELINE_MONTHS)
    visits: list[dict] = []

    for entry in COMPLETED_REGULAR:
        visits.append(
            build_regular_visit(
                entry["sortOrder"],
                regular_template,
                completed=True,
                target_odometer_km=entry["targetOdometerKm"],
                completed_odometer=entry["completedOdometer"],
                completed_at=entry.get("completedAt"),
                dealer=entry.get("dealer"),
                cost_uah=entry.get("costUah"),
                payments=entry.get("payments"),
                includes_oil_change=entry.get("includesOilChange", True),
                service_scope=entry.get("serviceScope", "oilService"),
                odometer_is_estimate=entry.get("odometerIsEstimate", False),
                cost_profile=entry.get("costProfile"),
                estimated_oil_portion_uah=entry.get("estimatedOilPortionUah"),
            )
        )

    last_oil_km = max(
        entry["completedOdometer"]
        for entry in COMPLETED_REGULAR
        if entry.get("includesOilChange")
    )
    max_milestone_km = max((milestone["targetOdometerKm"] for milestone in milestones), default=last_oil_km)
    windows = generate_interval_windows(last_oil_km, max_milestone_km)
    milestone_buckets = assign_milestones_to_windows(milestones, windows)

    for index, (window_from, target, window_to) in enumerate(windows):
        visits.append(
            build_interval_visit(
                window_from,
                target,
                window_to,
                regular_template,
                milestone_buckets[index],
            )
        )

    visits.sort(key=lambda item: item["sortOrder"])
    km_since_last_oil = odometer - last_oil_km
    next_window_from = last_oil_km + INTERVAL_KM_MIN
    next_window_to = last_oil_km + INTERVAL_KM_MAX
    next_target = last_oil_km + INTERVAL_KM_DEFAULT

    return {
        "schemaVersion": SCHEMA_VERSION,
        "vehicle": {
            "vin": "WVWZZZ3H5PE020759",
            "make": "Volkswagen",
            "model": "Arteon",
            "modelYear": 2024,
            "purchaseDate": PURCHASE_DATE,
            "engineLabel": "TDI",
            "odometerKm": odometer,
            "mileageBaselineKm": MILEAGE_BASELINE_KM,
            "mileageBaselineMonths": MILEAGE_BASELINE_MONTHS,
            "averageMonthlyMileageKm": avg_monthly,
        },
        "calculation": {
            "mode": "odometerWithEstimate",
            "primaryRule": "oilIntervalKm",
            "averageMonthlyMileageKm": avg_monthly,
            "averageMonthlyMileageUse": "dateEstimateOnly",
            "oilIntervalKm": {
                "min": INTERVAL_KM_MIN,
                "max": INTERVAL_KM_MAX,
                "default": INTERVAL_KM_DEFAULT,
            },
            "expectedOilChangeCostUah": EXPECTED_OIL_CHANGE_COST_UAH,
            "lastOilChangeOdometerKm": last_oil_km,
            "kmSinceLastOil": km_since_last_oil,
            "nextOilTargetKmMin": next_window_from,
            "nextOilTargetKmMax": next_window_to,
            "nextOilTargetKmDefault": next_target,
            "serviceRhythm": {
                "pattern": "oilIntervalKm",
                "intervalKmMin": INTERVAL_KM_MIN,
                "intervalKmMax": INTERVAL_KM_MAX,
                "intervalKmDefault": INTERVAL_KM_DEFAULT,
            },
            "regularService": {
                "intervalKmMin": INTERVAL_KM_MIN,
                "intervalKmMax": INTERVAL_KM_MAX,
                "intervalKmDefault": INTERVAL_KM_DEFAULT,
                "nextTargetKm": next_target,
                "nextWindowFromKm": next_window_from,
                "nextWindowToKm": next_window_to,
            },
        },
        "dealerStatement": build_dealer_statement(),
        "visits": visits,
    }


def import_upgrades(xlsx: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Доделки"]
    items: list[dict] = []
    for row in range(1, ws.max_row + 1):
        title = ws.cell(row, 1).value
        cost = ws.cell(row, 2).value
        if title and isinstance(title, str) and isinstance(cost, (int, float)):
            items.append(
                {
                    "id": f"upg-{len(items) + 1}",
                    "title": title.strip(),
                    "cost": int(cost),
                    "priority": len(items) + 1,
                    "status": "planned",
                }
            )
    return {"schemaVersion": 1, "items": items}


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    odometer = int(sys.argv[2]) if len(sys.argv) > 2 else CURRENT_ODOMETER
    xlsx_path = xlsx if xlsx.exists() else None
    if xlsx_path is None:
        print(f"Missing xlsx: {xlsx} — rebuilding plan from bundled milestones")

    plan = import_plan(xlsx_path, odometer)
    OUT_PLAN.write_text(json.dumps(plan, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    if xlsx_path is not None:
        upgrades = import_upgrades(xlsx_path)
        OUT_UPGRADES.write_text(json.dumps(upgrades, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    avg = plan["vehicle"]["averageMonthlyMileageKm"]
    calc = plan["calculation"]
    pending = [v for v in plan["visits"] if not v["isCompleted"]]
    next_visit = min(pending, key=lambda v: v["sortOrder"])
    print(f"Average monthly mileage: {avg} km (date estimates only)")
    print(f"Last oil @ {calc['lastOilChangeOdometerKm']} km; since last: {calc['kmSinceLastOil']} km")
    print(
        f"Next oil window: {calc['nextOilTargetKmMin']:,}–{calc['nextOilTargetKmMax']:,} km "
        f"(default {calc['nextOilTargetKmDefault']:,})".replace(",", ".")
    )
    print(f"Wrote {len(plan['visits'])} visits → {OUT_PLAN}")
    print(f"Next visit: {next_visit['title']}")


if __name__ == "__main__":
    main()
